from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

excel_file = '/home/twobe/Documentos/projeto1/meu_arquivo.xlsx'
wb = load_workbook(excel_file)
ws = wb.active

for cell in ws[1]:
    if cell.internal_value == 'Coluna1':
        col_index = cell.col_idx
        break

for cells_in_row in ws.iter_rows(min_row=0,
                                min_col=col_index,
                                max_col=col_index):
    print(cells_in_row)
    if cells_in_row[0].internal_value == 'Coluna2':
        cells_in_row[0].fill = PatternFill(patternType='solid', fgColor='e36464')

wb.save(excel_file)
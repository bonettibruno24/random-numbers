import pandas as pd
import os 
from collections import Counter
import sys

 # SAVED COMMIT ALL PROJECT PRE IMPLEMENTED LIB PYWIN32
while True:
    numbers = [33, 1, 20, 14, 31, 9, 22, 18, 29, 7, 28, 12, 35, 3, 26, 0, 32, 15, 19, 4, 21, 2, 25, 17, 34, 6, 27, 13, 36, 11, 30, 8, 23, 10,  5, 24, 16]
    last_moves = []
    path = os.path.join(os.getcwd(), "final.xlsx")
    # path = 'C:\Users\Maria\OneDrive\Área de Trabalho\project'

    def get_valid_numbers(prompt):
     while True:
        user_input = input(prompt).strip() 
        if user_input == "": 
            return None
        try:
            number = int(user_input)
            if number in numbers:
                return number
            else:
                print("Número inválido. Tente novamente.")
        except ValueError:
            print("Entrada inválida. Digite um número ou deixe em branco para continuar.")
            

    """
    Busca o último dígito de um número inteiro inserido, com valores e posição incluidos nas variavéis.
        Nos ternários eu verifico para pegar previus_position e later_position do numero.
    """            

    def get_last_digit(valor):
        if valor is None:  # Caso o valor seja None, retornamos um indicador ou mensagem
            return None
        if isinstance(valor, int):  # Verifica se o valor é um número inteiro
            return int(str(valor)[-1])
        else:
            raise ValueError("O valor fornecido não é um número válido.")
        

    def quick_sort(array):
        if len(array) <= 1:
            return array
        else:
            pivot = array[0]
            left = [x for x in array[1:] if x < pivot]
            right = [x for x in array[1:] if x >= pivot]
            return quick_sort(left) + [pivot] + quick_sort(right) 

    class GameAnalytics:
        def __init__(self):
            self.__sequence__numbers__ = []
            self.df_most_frequent = pd.DataFrame(columns=['Number', 'Frequency'])

        def process_fashion_to_dataframe(self, ordered_numbers):
            frequency = Counter(ordered_numbers)
            fashion = frequency.most_common()
            self.df_most_frequent = pd.DataFrame(fashion, columns=['Number', 'Frequency'])
            return self.df_most_frequent
            
        def last_play_analytics(self, prompt):
            last_play_analytics = input(prompt)
            try:
                self.__sequence__numbers__ = [int(num.strip()) for num in last_play_analytics.split(',')]
                last_digit = [get_last_digit(num) for num in self.__sequence__numbers__]
                ordered_numbers = quick_sort(last_digit)
                df_frequency_numbers = self.process_fashion_to_dataframe(ordered_numbers)
                    # print("chekando função aqui aqui ",df_frequency_numbers)
                return df_frequency_numbers
            except ValueError:
                print("Valor inválido. Tente novamente.")
                return self.last_play_analytics(prompt)

        def ask_init(self, command):
            result = input(command).strip().lower()     
            if result == "s":
                self.df_most_frequent = game.last_play_analytics("Cadastre a sequência numérica: ")
                return True  
            elif result == "n":
                excel_colum_g = pd.read_excel(path)
                ultiuma_jogada = excel_colum_g['Ultimas Jogadas'].values[-1]
                try:
                    if isinstance(ultiuma_jogada, str):
                        cleaned_data = ultiuma_jogada.replace('[', '').replace(']', '').strip()
                        sequence_numbers = [int(num.strip()) for num in cleaned_data.split(",")]
                        last_digit = [get_last_digit(num) for num in sequence_numbers]
                        self.__sequence__numbers__  = last_digit
                        print(f"Sequência numérica processada: {sequence_numbers}")
                        # print(f"Sequência numérica processada: {sequence_numbers}", type(sequence_numbers))
                        self.process_fashion_to_dataframe(last_digit)
                        # print("Sequência de números5555: ", self.df_most_frequent)
                        # print("Sequência de números: ", type(self.df_most_frequent))
                        return False
                    else:
                        raise ValueError("Formato inserido inválido.")
                except Exception as e:
                    print("Erro ao processar a sequência de números:", e)
                    return False

                print(ultiuma_jogada)
                print(type(ultiuma_jogada))
                return False
            else:
                print("Comando inválido. Tente novamente.")
                return self.ask_init(command)
    game = GameAnalytics()
    path = os.path.join(os.getcwd(), "final.xlsx")

    df_atualizado = pd.read_excel(path, sheet_name='Sheet1')
    # game.ask_init("Deseja continuar utilizando a mesma sequência anterior de 50 numeros? (S/N): ")
    game.ask_init("Deseja insir os ultimas rodadas? (S/N): ")
    # print(f"Sequência de números22222: {game.df_most_frequent}")
    # print(f"Sequência de números22222:", type(game.df_most_frequent))
   
    numero = get_valid_numbers("Digite um número: ")
    number1 = get_valid_numbers("Digite o segundo numero: ")
    number2 = get_valid_numbers("Digite o terceiro numero: ")
        
    print("sequencia numericao: ",game.__sequence__numbers__)
    last_digit_insert = get_last_digit(numero)
    last_digit_insert1 = get_last_digit(number1)
    last_digit_insert2= get_last_digit(number2)



    if number2 is not None:  # Verifica se number2 não é None antes de usar
        position_array_index = numbers.index(number2)

        previous_position = (
            numbers[position_array_index - 1] if position_array_index > 0 else numbers[-1]
        )
        later_position = (
            numbers[position_array_index + 1]
            if position_array_index < len(numbers) - 1
            else numbers[0]
        )
        algarism_previus = get_last_digit(previous_position)
        algarism_later = get_last_digit(later_position)

        # print(f"algarismo do anterior = {algarism_previus} & último algarismo = {algarism_later}")
        # print(f"posição no array {position_array_index}, número anterior: {previous_position}, número posterior: {later_position}")
        # print(f"n1= {last_digit_insert} & n2 = {last_digit_insert1} & n3 = {last_digit_insert2}")
    else:
        print("Nenhum número foi fornecido, cálculo ignorado.")

    # print("Tamanho do array Number:", len(df_most_frequent['Number']))
    # print("Tamanho do array Frequency:", len(df_most_frequent['Frequency']))

    number2 = number2 if number2 is not None and number2 != '' else None
    df = pd.DataFrame({
        'Valor 1': [numero],
        'Valor 2': [number1],
        'Valor 3': [number2],
        'Resultado' : [None],
        'Prob1': [None],
        'Prob2': [None],
        'Ultimas Jogadas': [game.df_most_frequent] 
        })
    if game.df_most_frequent is not None:
        df['Ultimas Jogadas'] = [game.__sequence__numbers__]
    # df = df.dropna(axis=1, how='all')


    # print("printando df: ", df)
    def veirify_algarism():
        current_set = {numero, number1, number2}
        required_set = {0, 26, 32}

        # teste = algarism_later or algarism_previus
        # print(teste)

        # print(f"Tipo de 'numero': {type(numero)}")
        # print(f"Tipo de 'number1': {type(number1)}")
        # print(f"Tipo de 'algarism_later': {type(algarism_later)}")
        # print(f"Tipo de 'algarism_previus': {type(algarism_previus)}")
        if  last_digit_insert2 == last_digit_insert and last_digit_insert1:
            print(last_digit_insert2, "OK")
            return True 
        elif numero or number1 == algarism_later or algarism_previus:
            print(f"{numero}, {number1} ,{number2} OK")
            return True
        elif last_digit_insert2 == algarism_previus and algarism_later:
            print(f"{numero}, {number1} ,{number2} OK")
            return True
        elif required_set == current_set:
            print(f"{numero}, {number1} ,{number2} OK")
            return True
        else:
            print(f"{numero}, {number1} ,{number2} X")
            return False

    # print("teste function",veirify_algarism())


    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Alignment

    # green = PatternFill("solid", fgColor="DDDDDD")
    # blue = PatternFill(fill_type='solid', start_color='00FF00', end_color='00FF00')

    # if veirify_algarism():
    #     WS['D2'] = "Ok"

    # else:
    #     WS['D2'] = "X"

    result_verify_algarism = veirify_algarism()

    if result_verify_algarism == True:
        df['Resultado'] = "OK"
    else:
        df['Resultado'] = "X"
        
    # print(df)
    # print(df_ler.columns)
    # print("col 1 ", df_ler['Valor 1'])
    # print("col 2 ",df_ler['Valor 2'])
    # print("col 3 ",df_ler['Valor 3'])

    """
    Prob1: Verifica quantas vezes o valor da variavel numero && numb1 está presente na lista dos ultimas 50 jogadas e divide por 50.

    Prob2: Verifica o num2 no excel quantas vezes ele saiu como OK na col 3 
    """

    def calculate_probability(df, numero, number1):

       # print(type(df))l  # Deve exibir: <class 'pandas.core.frame.DataFrame'>
        # print(f"Numero: {numero}, Number1: {number1}")
        # print(f"favorable_cases: ", game.df_most_frequent)
        favorable_cases = (
            df[df['Number'] == numero]['Frequency'].sum() +
            df[df['Number'] == number1]['Frequency'].sum()
        )
        total_operations = len(game.__sequence__numbers__)

        # print(f"favorable_cases: {favorable_cases}, {total_operations}")
        probability = (favorable_cases / total_operations) * 100 if total_operations > 0 else 0
        # print("printando probabilidade", probability)
        formated_probability = f"{probability:.2f}%"
        return formated_probability

    #Numbers contém os terminais de cada número
    probability = calculate_probability(game.df_most_frequent, numero, number1)
    if probability != "0.00%":
        df.at[0, 'Prob1'] = probability
    else:
        df['Prob1'] = None

    def calculate_probability_cell(df, number2):
        favorable_cases = len(df[(df['Valor 3'] == number2) & (df['Resultado'] == 'OK')])
        total_operations = len(df[(df['Valor 3'] == number2)])
        probability = (favorable_cases / total_operations) * 100 if total_operations > 0 else 0
        formated_probability = f"{probability:.2f}%"
        # print(f"favorable_cases: {favorable_cases}, {total_operations}")
        return formated_probability

    
    if getattr(sys, 'frozen', False):
        # Se o script está sendo executado a partir de um executável
        caminho_final = os.path.join(sys._MEIPASS, 'final.xlsx')
    else:
        # Se o script está sendo executado no ambiente de desenvolvimento
        caminho_final = os.path.join(os.getcwd(), 'final.xlsx')

    print(f"O caminho para o arquivo é: {caminho_final}")

    # Lendo o arquivo Excel
    # print(df_atualizado)
    probability_two = calculate_probability_cell(df_atualizado, number2)
    print(f"Probabilidade1: {probability}%")
    print(f"Probabilidade2: {probability_two}")
    occurrences = len(df_atualizado[df_atualizado['Valor 3'] == number2])
    print(f"O número {number2} apareceu {occurrences} vezes na coluna 'Valor 3'.")
    df.at[0, 'Prob2'] = probability_two

    # print(df_atualizado)
        # print(f"Probabilidade2: {probability:.2f}%")

    if os.path.exists(path):
        df_existente = pd.read_excel(caminho_final, sheet_name='Sheet1')
        df_atualizado = pd.concat([df_existente, df ], ignore_index=True)
        with pd.ExcelWriter(path, engine='openpyxl', mode= 'a', if_sheet_exists='replace') as writer:
            df_atualizado.to_excel(writer, sheet_name='Sheet1', index=False)
            print(f"Dados inseridos com sucesso! {path}")
    else:
        df.to_excel(path, index=False)
        print(f"Dados inseridos com sucesso! {path}")

    wb = load_workbook(path)
    ws = wb.active
    wb.save(path)

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for cell in ws[1]:
        if cell.value == 'Resultado':
            col_index = cell.col_idx
            break

    for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value == 'OK':
                cell.fill = green_fill
            elif cell.value == 'X':
                cell.fill = red_fill
    wb.save(path)